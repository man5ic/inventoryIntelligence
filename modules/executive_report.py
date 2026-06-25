"""
Phase 8: Filtered Executive Report Generator
Generates a professional filtered dashboard Excel export.
Reuses all existing analytics — no duplicate calculations.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import Optional

# ── Shared palette ────────────────────────────────────────────────────────────
COLORS = {
    'navy':     '1E3A5F',
    'white':    'FFFFFF',
    'blue':     '2563EB',
    'sky':      'DBEAFE',
    'green_bg': 'DCFCE7',
    'green_fg': '166534',
    'red_bg':   'FEE2E2',
    'red_fg':   '991B1B',
    'amber_bg': 'FEF3C7',
    'amber_fg': '92400E',
    'slate_bg': 'F8FAFC',
    'slate':    '64748B',
    'border':   'CBD5E1',
    'A_class':  'FEF3C7',
    'B_class':  'DBEAFE',
    'C_class':  'F0FDF4',
    'critical': 'FEE2E2',
    'medium':   'FEF3C7',
}

thin   = Side(border_style='thin', color=COLORS['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Style helpers ─────────────────────────────────────────────────────────────

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _section_header(ws, row: int, cols: int, text: str,
                    bg: str = None, fg: str = 'FFFFFF', size: int = 11):
    bg = bg or COLORS['navy']
    _merge(ws, row, 1, row, cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 24


def _col_header(ws, row: int, headers: list, start_col: int = 1):
    for i, h in enumerate(headers, start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = PatternFill('solid', fgColor=COLORS['blue'])
        c.font = Font(bold=True, color='FFFFFF', size=9)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def _kv(ws, row, col, label, value, label_bg='F1F5F9', value_bg='FFFFFF'):
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = Font(bold=True, size=9, color=COLORS['navy'])
    lc.fill = PatternFill('solid', fgColor=label_bg)
    lc.border = BORDER
    lc.alignment = Alignment(vertical='center')

    vc = ws.cell(row=row, column=col + 1, value=value)
    vc.font = Font(size=9, color='1E293B')
    vc.fill = PatternFill('solid', fgColor=value_bg)
    vc.border = BORDER
    vc.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 18


def _auto_width(ws, min_w=12, max_w=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[letter].width = min(max(w + 3, min_w), max_w)


# ─────────────────────────────────────────────────────────────────────────────
# Title Block
# ─────────────────────────────────────────────────────────────────────────────

def _title_block(ws, filter_label: str, cols: int = 8):
    ws.sheet_view.showGridLines = False

    # Title row
    _merge(ws, 1, 1, 1, cols)
    t = ws.cell(row=1, column=1,
                value='INVENTORY INTELLIGENCE SYSTEM  |  EXECUTIVE REPORT')
    t.font  = Font(bold=True, size=14, color=COLORS['white'])
    t.fill  = PatternFill('solid', fgColor=COLORS['navy'])
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle / filter context
    _merge(ws, 2, 1, 2, cols - 2)
    s = ws.cell(row=2, column=1, value=f'Filter: {filter_label}')
    s.font  = Font(bold=True, size=10, color=COLORS['navy'])
    s.fill  = PatternFill('solid', fgColor='EFF6FF')
    s.alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 20

    _merge(ws, 2, cols - 1, 2, cols)
    d = ws.cell(row=2, column=cols - 1,
                value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}")
    d.font = Font(italic=True, size=9, color=COLORS['slate'])
    d.alignment = Alignment(horizontal='right', vertical='center')
    d.fill = PatternFill('solid', fgColor='EFF6FF')

    ws.row_dimensions[3].height = 8   # spacer


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Executive Summary + KPIs
# ─────────────────────────────────────────────────────────────────────────────

def _write_exec_summary_sheet(wb, inv_result, forecast_result, exec_summary,
                               alert_result, filter_label, filter_ctx=None):
    ws = wb.create_sheet('Executive Summary')
    COLS = 6
    _title_block(ws, filter_label, cols=COLS)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 20

    row = 4

    # ── Headline ──
    headline = exec_summary.get('headline', '') if exec_summary else ''
    if headline:
        _merge(ws, row, 1, row, COLS)
        hc = ws.cell(row=row, column=1, value=headline)
        hc.font  = Font(bold=True, size=11, color='1E40AF')
        hc.fill  = PatternFill('solid', fgColor=COLORS['sky'])
        hc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 36
        row += 2

    # ── KPI Cards (2-column grid) ──
    _section_header(ws, row, COLS, '📊  KEY PERFORMANCE INDICATORS', bg='1D4ED8')
    row += 1

    kpis = exec_summary.get('kpis', {}) if exec_summary else {}
    inv_sum = inv_result.get('summary', {}) if inv_result else {}
    fc_sum  = forecast_result.get('summary', {}) if forecast_result else {}
    alrt    = alert_result if alert_result else {}

    kpi_data = [
        ('Total Materials',       f"{kpis.get('total_items', inv_sum.get('total_items', 0)):,}"),
        ('Total Inventory Value', f"₹{kpis.get('total_value', inv_sum.get('total_value', 0)):,.0f}"),
        ('Class A Materials',     f"{kpis.get('a_items', inv_sum.get('a_items', 0)):,}"),
        ('Class B Materials',     f"{inv_sum.get('b_items', 0):,}"),
        ('Class C Materials',     f"{inv_sum.get('c_items', 0):,}"),
        ('Peak Demand Materials', f"{kpis.get('peak_materials', fc_sum.get('peak_count', 0)):,}"),
        ('Next Quarter Cost Est.',f"₹{kpis.get('next_qtr_cost', fc_sum.get('next_quarter_cost', 0)):,.0f}"),
        ('Cost Change Forecast',  f"{kpis.get('cost_change_pct', fc_sum.get('avg_pct_change', 0)):+.1f}%"),
        ('Active Alerts',         f"{alrt.get('total', 0):,}"),
        ('Critical Alerts',       f"{alrt.get('critical_count', 0):,}"),
    ]

    bg_cycle = ['EFF6FF', 'F0F9FF', 'F0FDF4', 'FFFBEB', 'FFF1F2']
    for i, (label, val) in enumerate(kpi_data):
        col_pair = (i % 3) * 2 + 1   # 3 columns across
        r = row + i // 3
        bg = bg_cycle[i % len(bg_cycle)]
        lc = ws.cell(row=r, column=col_pair, value=label)
        lc.font  = Font(bold=True, size=9, color=COLORS['navy'])
        lc.fill  = PatternFill('solid', fgColor=bg)
        lc.border = BORDER
        lc.alignment = Alignment(vertical='center')
        vc = ws.cell(row=r, column=col_pair + 1, value=val)
        vc.font  = Font(bold=True, size=10, color=COLORS['blue'])
        vc.fill  = PatternFill('solid', fgColor='FFFFFF')
        vc.border = BORDER
        vc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22

    row += (len(kpi_data) + 2) // 3 + 2

    # ── Business Summary ──
    if exec_summary and exec_summary.get('bullets'):
        _section_header(ws, row, COLS, '📋  BUSINESS SUMMARY',
                        bg='0F172A', size=10)
        row += 1
        for bullet in exec_summary.get('bullets', []):
            _merge(ws, row, 1, row, COLS)
            bc = ws.cell(row=row, column=1, value=f'• {bullet}')
            bc.font = Font(size=9, color='1E293B')
            bc.fill = PatternFill('solid', fgColor=COLORS['slate_bg'] if row % 2 == 0 else COLORS['white'])
            bc.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1

    # ── Risk Flags ──
    if exec_summary and exec_summary.get('risk_flags'):
        _section_header(ws, row, COLS, '⚠️  RISK FLAGS', bg='991B1B', size=10)
        row += 1
        for flag in exec_summary.get('risk_flags', []):
            _merge(ws, row, 1, row, COLS)
            fc_cell = ws.cell(row=row, column=1, value=f'⚠ {flag}')
            fc_cell.font = Font(size=9, color=COLORS['red_fg'])
            fc_cell.fill = PatternFill('solid', fgColor=COLORS['red_bg'])
            fc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1

    # ── Recommendations ──
    if exec_summary and exec_summary.get('recommendations'):
        _section_header(ws, row, COLS, '✅  RECOMMENDED ACTIONS', bg='166534', size=10)
        row += 1
        for rec in exec_summary.get('recommendations', []):
            _merge(ws, row, 1, row, COLS)
            rc_cell = ws.cell(row=row, column=1, value=f'✓ {rec}')
            rc_cell.font = Font(size=9, color=COLORS['green_fg'])
            rc_cell.fill = PatternFill('solid', fgColor=COLORS['green_bg'])
            rc_cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 28
            row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Procurement Priority Ranking
# ─────────────────────────────────────────────────────────────────────────────

def _write_ranking_sheet(wb, ranking_result, filter_label):
    ws = wb.create_sheet('Procurement Ranking')
    COLS = 7
    _title_block(ws, filter_label, cols=COLS)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 20

    row = 4
    _section_header(ws, row, COLS, '🏆  PROCUREMENT PRIORITY RANKING — TOP MATERIALS',
                    bg='1D4ED8')
    row += 1
    _col_header(ws, row, ['Material Code', 'Description', 'ABC Class',
                           'Priority Score', 'Priority Level',
                           'Annual Value (₹)', 'Recommended Action'])
    row += 1

    priority_bg = {'CRITICAL': COLORS['red_bg'],   'HIGH': COLORS['amber_bg'],
                   'MEDIUM':   COLORS['sky'],       'LOW': COLORS['green_bg'],
                   'Critical': COLORS['red_bg'],   'Monitor': COLORS['amber_bg'],
                   'Stable':   COLORS['green_bg']}
    priority_fg = {'CRITICAL': COLORS['red_fg'],    'HIGH': COLORS['amber_fg'],
                   'MEDIUM':   COLORS['blue'],      'LOW': COLORS['green_fg'],
                   'Critical': COLORS['red_fg'],   'Monitor': COLORS['amber_fg'],
                   'Stable':   COLORS['green_fg']}
    abc_bg = {'A': COLORS['A_class'], 'B': COLORS['B_class'], 'C': COLORS['C_class']}

    # ranking_result may be a dict with 'rankings' key or a list
    if isinstance(ranking_result, dict):
        items = ranking_result.get('rankings', [])
    else:
        items = ranking_result or []

    for item in items:
        lvl = item.get('priority_level', 'LOW')
        row_bg = COLORS['slate_bg'] if row % 2 == 0 else COLORS['white']
        vals = [
            item.get('material_code', ''),
            item.get('description', ''),
            item.get('abc_class', ''),
            f"{item.get('priority_score', 0):.1f}",
            lvl,
            f"₹{item.get('annual_value', 0):,.0f}",
            item.get('recommendation', item.get('action', '')),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 3:
                c.fill = PatternFill('solid', fgColor=abc_bg.get(str(val).upper(), COLORS['white']))
            elif col == 5:
                c.fill = PatternFill('solid', fgColor=priority_bg.get(lvl, COLORS['white']))
                c.font = Font(bold=True, size=9, color=priority_fg.get(lvl, '000000'))
            else:
                c.fill = PatternFill('solid', fgColor=row_bg)
                c.font = Font(size=9)
            c.border = BORDER
            c.alignment = Alignment(
                horizontal='center' if col in (3, 4, 5) else 'right' if col == 6 else 'left',
                vertical='center', wrap_text=(col == 7)
            )
        ws.row_dimensions[row].height = 22
        row += 1

    ws.freeze_panes = 'A7'


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Top Risk Materials
# ─────────────────────────────────────────────────────────────────────────────

def _write_risk_materials_sheet(wb, risk_result, filter_label):
    ws = wb.create_sheet('Top Risk Materials')
    COLS = 6
    _title_block(ws, filter_label, cols=COLS)

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 40

    row = 4
    _section_header(ws, row, COLS, '🔴  TOP RISK MATERIALS — REQUIRES ATTENTION', bg='991B1B')
    row += 1
    _col_header(ws, row, ['Material Code', 'Description', 'ABC Class',
                           'Risk Score', 'Forecast Status', 'Risk Insight'])
    row += 1

    status_bg = {'PEAK': COLORS['green_bg'], 'NORMAL': COLORS['sky'],
                 'LOW': COLORS['amber_bg']}

    # risk_result may be a dict with 'risks' key or a list
    if isinstance(risk_result, dict):
        items = risk_result.get('risks', [])
    else:
        items = risk_result or []

    for item in items:
        row_bg = COLORS['red_bg'] if row % 2 == 0 else 'FFF5F5'
        status = item.get('forecast_status', item.get('trend', 'NORMAL'))
        # Build insight from alerts list if available
        insight_text = item.get('risk_insight', '')
        if not insight_text:
            alerts_list = item.get('alerts', [])
            insight_text = '; '.join(alerts_list[:2]) if alerts_list else item.get('insight', '')

        vals = [
            item.get('material_code', ''),
            item.get('description', ''),
            item.get('abc_class', ''),
            f"{item.get('risk_score', 0):.1f}",
            status,
            insight_text,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 5:
                c.fill = PatternFill('solid', fgColor=status_bg.get(status, COLORS['white']))
                c.font = Font(bold=True, size=9)
            else:
                c.fill = PatternFill('solid', fgColor=row_bg)
                c.font = Font(size=9)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center' if col in (3, 4, 5) else 'left',
                                    vertical='top', wrap_text=(col == 6))
        ws.row_dimensions[row].height = 36
        row += 1

    ws.freeze_panes = 'A7'


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Smart Insights
# ─────────────────────────────────────────────────────────────────────────────

def _write_insights_sheet(wb, exec_summary, forecast_result, filter_label):
    ws = wb.create_sheet('Smart Insights')
    COLS = 5
    _title_block(ws, filter_label, cols=COLS)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    row = 4
    _section_header(ws, row, COLS, '💡  AI-GENERATED SMART INSIGHTS', bg='1D4ED8')
    row += 1

    # Procurement insights
    proc_insights = exec_summary.get('procurement_insights', []) if exec_summary else []
    if proc_insights:
        _section_header(ws, row, COLS, 'PROCUREMENT & SUPPLY CHAIN INSIGHTS',
                        bg='374151', size=9)
        row += 1
        for i, insight in enumerate(proc_insights, 1):
            _merge(ws, row, 1, row, COLS)
            ic = ws.cell(row=row, column=1,
                         value=f'{i:02d}.  {insight}')
            ic.font = Font(size=9, color='1E293B')
            ic.fill = PatternFill('solid', fgColor=COLORS['sky'] if i % 2 else COLORS['white'])
            ic.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 32
            row += 1
        row += 1

    # Forecast exec insights
    fc_insights = forecast_result.get('exec_insights', []) if forecast_result else []
    if fc_insights:
        _section_header(ws, row, COLS, 'FORECAST & DEMAND INSIGHTS',
                        bg='374151', size=9)
        row += 1
        for i, insight in enumerate(fc_insights, 1):
            _merge(ws, row, 1, row, COLS)
            ic = ws.cell(row=row, column=1, value=f'{i:02d}.  {insight}')
            ic.font = Font(size=9, color='1E293B')
            ic.fill = PatternFill('solid',
                                  fgColor=COLORS['green_bg'] if i % 2 else COLORS['white'])
            ic.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 32
            row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Forecast Summary
# ─────────────────────────────────────────────────────────────────────────────

def _write_forecast_summary_sheet(wb, forecast_result, filter_label):
    ws = wb.create_sheet('Forecast Summary')
    COLS = 8
    _title_block(ws, filter_label, cols=COLS)

    for col, w in zip('ABCDEFGH', [16, 24, 14, 14, 14, 16, 16, 40]):
        ws.column_dimensions[col].width = w

    row = 4
    _section_header(ws, row, COLS, '📈  FORECAST SUMMARY — ALL MATERIALS', bg='1D4ED8')
    row += 1

    summary = forecast_result.get('summary', {}) if forecast_result else {}
    kv_pairs = [
        ('Total Forecasted Materials', f"{summary.get('total_forecasted', 0):,}"),
        ('Peak Demand Materials',      f"{summary.get('peak_count', 0):,}"),
        ('Low Demand Materials',       f"{summary.get('low_count', 0):,}"),
        ('Normal Demand Materials',    f"{summary.get('normal_count', 0):,}"),
        ('Next Quarter Cost Est.',     f"₹{summary.get('next_quarter_cost', 0):,.0f}"),
        ('Avg Demand Change %',        f"{summary.get('avg_pct_change', 0):+.1f}%"),
    ]
    for i, (label, val) in enumerate(kv_pairs):
        col_off = (i % 3) * 2 + 1
        r = row + i // 3
        lc = ws.cell(row=r, column=col_off, value=label)
        lc.font = Font(bold=True, size=9, color=COLORS['navy'])
        lc.fill = PatternFill('solid', fgColor='EFF6FF')
        lc.border = BORDER
        vc = ws.cell(row=r, column=col_off + 1, value=val)
        vc.font = Font(bold=True, size=10, color=COLORS['blue'])
        vc.fill = PatternFill('solid', fgColor=COLORS['white'])
        vc.border = BORDER
        vc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 22

    row += (len(kv_pairs) + 2) // 3 + 2

    # Material forecast table
    forecasts = forecast_result.get('forecasts', []) if forecast_result else []
    if forecasts:
        _section_header(ws, row, COLS, 'MATERIAL-LEVEL FORECAST DETAIL', bg='374151', size=9)
        row += 1
        _col_header(ws, row, ['Code', 'Description', 'Status', 'Change %',
                               'Next Month', 'Next Quarter', 'Next Year', 'Insight'])
        row += 1

        status_bg = {'PEAK': COLORS['green_bg'], 'NORMAL': COLORS['sky'],
                     'LOW': COLORS['amber_bg']}
        status_fg = {'PEAK': COLORS['green_fg'], 'NORMAL': COLORS['blue'],
                     'LOW': COLORS['amber_fg']}

        for f in forecasts:
            status = f.get('forecast_status', 'NORMAL')
            row_bg = COLORS['slate_bg'] if row % 2 == 0 else COLORS['white']
            vals = [
                f.get('material_code', ''),
                f.get('description', ''),
                status,
                f"{f.get('pct_change', 0):+.1f}%",
                f"{f.get('next_month_qty', 0):,.0f}",
                f"{f.get('next_quarter_qty', 0):,.0f}",
                f"{f.get('next_year_qty', 0):,.0f}",
                f.get('insight', ''),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                if col == 3:
                    c.fill = PatternFill('solid', fgColor=status_bg.get(status, COLORS['white']))
                    c.font = Font(bold=True, size=9, color=status_fg.get(status, '000000'))
                else:
                    c.fill = PatternFill('solid', fgColor=row_bg)
                    c.font = Font(size=9)
                c.border = BORDER
                c.alignment = Alignment(
                    horizontal='center' if col in (3, 4, 5, 6, 7) else 'left',
                    vertical='top', wrap_text=(col == 8)
                )
            ws.row_dimensions[row].height = 28
            row += 1

    ws.freeze_panes = f'A{row - len(forecasts)}'


# ─────────────────────────────────────────────────────────────────────────────
# Main: Generate Filtered Executive Report
# ─────────────────────────────────────────────────────────────────────────────

def generate_executive_report(inv_result, forecast_result, alert_result,
                               exec_summary, ranking_result, risk_result,
                               plant: str = '__all__', region: str = '__all__',
                               period: str = '__all__') -> BytesIO:
    """
    Build the filtered executive Excel report.
    All inputs are pre-computed pipeline outputs — zero duplicate calculation.
    """
    # Build human-readable filter label
    parts = []
    if plant and plant != '__all__':
        parts.append(f'Plant: {plant}')
    if region and region != '__all__':
        parts.append(f'Region: {region}')
    if period and period != '__all__':
        parts.append(f'Period: {period}')
    filter_label = '  |  '.join(parts) if parts else 'All Plants · All Regions · Full Period'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_exec_summary_sheet(wb, inv_result, forecast_result,
                               exec_summary, alert_result, filter_label)
    _write_ranking_sheet(wb, ranking_result, filter_label)
    _write_risk_materials_sheet(wb, risk_result, filter_label)
    _write_insights_sheet(wb, exec_summary, forecast_result, filter_label)
    if forecast_result:
        _write_forecast_summary_sheet(wb, forecast_result, filter_label)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
