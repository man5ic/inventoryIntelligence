"""
Phase 8: Material Report Generation Engine
Generates professional Excel reports for individual materials.
Reuses existing material_intelligence outputs — no duplicate calculations.
"""
import base64
import io
from io import BytesIO
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── shared palette (mirrors report_export.py) ──────────────────────────────
COLORS = {
    'navy':      '1E3A5F',
    'white':     'FFFFFF',
    'blue':      '2563EB',
    'sky':       'DBEAFE',
    'green_bg':  'DCFCE7',
    'green_fg':  '166534',
    'red_bg':    'FEE2E2',
    'red_fg':    '991B1B',
    'amber_bg':  'FEF3C7',
    'amber_fg':  '92400E',
    'slate_bg':  'F8FAFC',
    'slate':     '64748B',
    'border':    'CBD5E1',
    'A_class':   'FEF3C7',
    'B_class':   'DBEAFE',
    'C_class':   'F0FDF4',
}

thin = Side(border_style='thin', color=COLORS['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Style helpers ────────────────────────────────────────────────────────────

def _merge_header(ws, row: int, start_col: int, end_col: int,
                  text: str, bg: str, fg: str = 'FFFFFF', size: int = 11):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = Font(bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22


def _kv_row(ws, row: int, label: str, value, label_col=1, value_col=2,
            bg_label='F1F5F9', bg_value='FFFFFF', bold_label=True):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = Font(bold=bold_label, size=9, color=COLORS['navy'])
    lc.fill = PatternFill('solid', fgColor=bg_label)
    lc.border = BORDER
    lc.alignment = Alignment(vertical='center')

    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font = Font(size=9, color='1E293B')
    vc.fill = PatternFill('solid', fgColor=bg_value)
    vc.border = BORDER
    vc.alignment = Alignment(vertical='center')
    ws.row_dimensions[row].height = 18


def _section_title(ws, row: int, cols: int, text: str, icon: str = '',
                   bg: str = None, fg: str = None):
    bg = bg or COLORS['navy']
    fg = fg or COLORS['white']
    full_text = f'{icon}  {text}' if icon else text
    _merge_header(ws, row, 1, cols, full_text, bg, fg, size=10)


def _col_header(ws, row: int, headers: list):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = PatternFill('solid', fgColor=COLORS['blue'])
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 26


def _auto_width(ws, min_w=12, max_w=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[letter].width = min(max(w + 3, min_w), max_w)


# ── Title block ─────────────────────────────────────────────────────────────

def _write_title_block(ws, material_data: dict, cols: int = 6):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # Row 1 — big title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    cell = ws.cell(row=1, column=1,
                   value='MATERIAL INTELLIGENCE REPORT  |  INVENTORY INTELLIGENCE SYSTEM')
    cell.font = Font(bold=True, size=13, color=COLORS['white'])
    cell.fill = PatternFill('solid', fgColor=COLORS['navy'])
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Row 2 — material name + generated date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols - 2)
    ws.cell(row=2, column=1,
            value=f"{material_data.get('material_code', '')}  –  {material_data.get('description', '')}").font = \
        Font(bold=True, size=11, color=COLORS['navy'])
    ws.cell(row=2, column=1).alignment = Alignment(vertical='center')
    ws.row_dimensions[2].height = 22

    ws.merge_cells(start_row=2, start_column=cols - 1, end_row=2, end_column=cols)
    ts_cell = ws.cell(row=2, column=cols - 1,
                      value=f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}")
    ts_cell.font = Font(italic=True, size=9, color=COLORS['slate'])
    ts_cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.row_dimensions[3].height = 6   # spacer


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Material Overview
# ─────────────────────────────────────────────────────────────────────────────

def _write_overview_sheet(wb, material_data: dict):
    ws = wb.create_sheet('Material Overview')
    ws.sheet_view.showGridLines = False
    COLS = 4

    _write_title_block(ws, material_data, cols=COLS)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22

    row = 4
    _section_title(ws, row, COLS, 'MATERIAL OVERVIEW', '📦')
    row += 1

    # Health Score banner
    health = material_data.get('health', {})
    score = health.get('score', 0)
    label = health.get('label', 'Monitor')
    health_bg = {'Healthy': COLORS['green_bg'], 'Monitor': COLORS['amber_bg'],
                 'At Risk': 'FECACA', 'Critical': COLORS['red_bg']}.get(label, COLORS['sky'])
    health_fg = {'Healthy': COLORS['green_fg'], 'Monitor': COLORS['amber_fg'],
                 'At Risk': COLORS['red_fg'], 'Critical': COLORS['red_fg']}.get(label, COLORS['navy'])

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    hc = ws.cell(row=row, column=1,
                 value=f'Material Health Score: {score}/100  –  {label}')
    hc.font = Font(bold=True, size=12, color=health_fg)
    hc.fill = PatternFill('solid', fgColor=health_bg)
    hc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 24
    row += 1

    # Two-column KV grid
    kv_pairs_left = [
        ('Material Code',   material_data.get('material_code', '')),
        ('Description',     material_data.get('description', '')),
        ('Category',        material_data.get('category', '')),
        ('ABC Class',       material_data.get('abc_class', '')),
        ('Inventory Status',material_data.get('inventory_status', '')),
        ('Unit Price (₹)',  f"₹{material_data.get('unit_price', 0):,.2f}"),
        ('Annual Value (₹)',f"₹{material_data.get('annual_value', 0):,.0f}"),
    ]
    kv_pairs_right = [
        ('Total Demand',     f"{material_data.get('total_demand', 0):,.0f} units"),
        ('Forecast Demand',  f"{material_data.get('forecast_demand', 0):,.0f} units"),
        ('Safety Stock',     f"{material_data.get('safety_stock', 0):,.1f} units"),
        ('Reorder Point',    f"{material_data.get('reorder_point', 0):,.1f} units"),
        ('EOQ',              f"{material_data.get('eoq', 0):,.1f} units"),
        ('Lead Time',        f"{material_data.get('lead_time', 0):.0f} days"),
        ('Avg Monthly Demand', f"{material_data.get('avg_monthly_demand', 0):,.1f} units"),
    ]

    for (ll, lv), (rl, rv) in zip(kv_pairs_left, kv_pairs_right):
        _kv_row(ws, row, ll, lv, label_col=1, value_col=2)
        _kv_row(ws, row, rl, rv, label_col=3, value_col=4)
        row += 1

    row += 1
    _section_title(ws, row, COLS, 'RECOMMENDATION & STATUS', '💡',
                   bg=COLORS['blue'], fg=COLORS['white'])
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    rc = ws.cell(row=row, column=1, value=material_data.get('recommendation', ''))
    rc.font = Font(size=10, bold=True, color=COLORS['navy'])
    rc.fill = PatternFill('solid', fgColor=COLORS['sky'])
    rc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[row].height = 36
    row += 1

    # Health breakdown
    row += 1
    _section_title(ws, row, COLS, 'HEALTH SCORE BREAKDOWN', '🏥',
                   bg='374151', fg=COLORS['white'])
    row += 1
    breakdown = health.get('breakdown', {})
    breakdown_labels = {
        'demand_stability': ('Demand Stability', 30),
        'forecast_trend':   ('Forecast Trend',   25),
        'procurement_risk': ('Procurement Risk',  25),
        'inventory_safety': ('Inventory Safety',  20),
    }
    for key, (label_str, max_pts) in breakdown_labels.items():
        pts = breakdown.get(key, 0)
        pct = round(pts / max_pts * 100)
        bar = '█' * int(pct / 10) + '░' * (10 - int(pct / 10))
        _kv_row(ws, row, f'{label_str} (max {max_pts})',
                f'{pts} pts  [{bar}]  {pct}%',
                label_col=1, value_col=2)
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Forecast Section
# ─────────────────────────────────────────────────────────────────────────────

def _write_forecast_sheet(wb, material_data: dict):
    ws = wb.create_sheet('Forecast Analysis')
    ws.sheet_view.showGridLines = False
    COLS = 4
    _write_title_block(ws, material_data, cols=COLS)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 22

    row = 4
    _section_title(ws, row, COLS, 'FORECAST SUMMARY', '📈')
    row += 1

    pct = material_data.get('pct_change', 0)
    status = material_data.get('forecast_status', 'NORMAL')
    status_bg = {'PEAK': COLORS['green_bg'], 'NORMAL': COLORS['sky'],
                 'LOW': COLORS['amber_bg']}.get(status, COLORS['sky'])

    # Forecast Status banner
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    fc = ws.cell(row=row, column=1,
                 value=f'Forecast Status: {status}  |  Demand Change: {pct:+.1f}%  |  Model: {material_data.get("best_model","").replace("_"," ").title()}')
    fc.font = Font(bold=True, size=10, color=COLORS['navy'])
    fc.fill = PatternFill('solid', fgColor=status_bg)
    fc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 22
    row += 1

    monthly = material_data.get('monthly_forecast', [])
    next_month = monthly[0]['forecast_qty'] if monthly else 0
    next_qtr   = sum(m['forecast_qty'] for m in monthly[:3]) if len(monthly) >= 3 else 0
    next_yr    = sum(m['forecast_qty'] for m in monthly[:12]) if len(monthly) >= 12 else 0

    kv_left = [
        ('Next Month Demand',   f"{next_month:,.1f} units"),
        ('Next Quarter Demand', f"{next_qtr:,.1f} units"),
        ('Next Year Demand',    f"{next_yr:,.1f} units"),
        ('Demand Change %',     f"{pct:+.1f}%"),
    ]
    kv_right = [
        ('Forecast Status',     status),
        ('Forecast Trend',      'Upward ▲' if pct > 5 else ('Downward ▼' if pct < -5 else 'Stable →')),
        ('Demand Stability',    material_data.get('demand_stability', 'Unknown')),
        ('Forecast Insight',    material_data.get('insight', '')),
    ]
    for (ll, lv), (rl, rv) in zip(kv_left, kv_right):
        _kv_row(ws, row, ll, lv, label_col=1, value_col=2)
        _kv_row(ws, row, rl, rv, label_col=3, value_col=4)
        row += 1

    # Monthly forecast table
    row += 1
    _section_title(ws, row, COLS, 'MONTHLY FORECAST DETAIL', '📅',
                   bg=COLORS['blue'], fg=COLORS['white'])
    row += 1
    _col_header(ws, row, ['Month', 'Forecast Qty', 'Forecast Cost (₹)', 'Trend'])
    row += 1

    prev_qty = None
    for m in monthly[:12]:
        qty  = m.get('forecast_qty', 0)
        cost = m.get('forecast_cost', 0)
        trend = ''
        if prev_qty is not None:
            trend = '▲' if qty > prev_qty * 1.02 else ('▼' if qty < prev_qty * 0.98 else '→')
        prev_qty = qty
        bg = COLORS['slate_bg'] if row % 2 == 0 else COLORS['white']
        for col, val in enumerate([m.get('month', ''), f"{qty:,.1f}", f"₹{cost:,.0f}", trend], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center' if col in (1, 3, 4) else 'right')
        row += 1

    # Embed chart image if available
    chart_data = material_data.get('chart_data', {})
    forecast_img_b64 = chart_data.get('forecast_chart_b64') or chart_data.get('line_chart_b64')
    if forecast_img_b64:
        try:
            img_bytes = base64.b64decode(forecast_img_b64)
            img_stream = BytesIO(img_bytes)
            img = XLImage(img_stream)
            img.width  = 520
            img.height = 260
            ws.add_image(img, f'A{row + 2}')
            # Reserve rows for the image
            for r in range(row + 2, row + 18):
                ws.row_dimensions[r].height = 18
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Procurement Section
# ─────────────────────────────────────────────────────────────────────────────

def _write_procurement_sheet(wb, material_data: dict):
    ws = wb.create_sheet('Procurement Plan')
    ws.sheet_view.showGridLines = False
    COLS = 4
    _write_title_block(ws, material_data, cols=COLS)
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 24

    row = 4
    _section_title(ws, row, COLS, 'PROCUREMENT RECOMMENDATIONS', '🛒')
    row += 1

    orders = material_data.get('orders', {})
    procurement = material_data.get('procurement', {})

    kv_left = [
        ('Weekly Order Quantity',     f"{orders.get('weekly_qty', orders.get('weekly_order_qty', 0)):,.1f} units"),
        ('Monthly Order Quantity',    f"{orders.get('monthly_qty', orders.get('monthly_order_qty', 0)):,.1f} units"),
        ('Quarterly Order Quantity',  f"{orders.get('eoq_qty', orders.get('quarterly_order_qty', 0)):,.1f} units"),
        ('EOQ (Economic Order Qty)',  f"{material_data.get('eoq', 0):,.1f} units"),
        ('Safety Stock Required',     f"{material_data.get('safety_stock', 0):,.1f} units"),
    ]
    kv_right = [
        ('Reorder Point',             f"{material_data.get('reorder_point', 0):,.1f} units"),
        ('Lead Time',                 f"{material_data.get('lead_time', 0):.0f} days"),
        ('Order Frequency',           f"Every {orders.get('eoq_order_interval_days', 30)} days"),
        ('Procurement Urgency',       orders.get('urgency', 'Standard')),
        ('Basis',                     orders.get('basis', '')),
    ]
    for (ll, lv), (rl, rv) in zip(kv_left, kv_right):
        _kv_row(ws, row, ll, lv, label_col=1, value_col=2)
        _kv_row(ws, row, rl, rv, label_col=3, value_col=4)
        row += 1

    row += 1
    _section_title(ws, row, COLS, 'FUTURE PROCUREMENT ESTIMATION', '📊',
                   bg=COLORS['blue'], fg=COLORS['white'])
    row += 1
    _col_header(ws, row, ['Period', 'Est. Quantity', 'Est. Cost (₹)', 'Notes'])
    row += 1

    # Build estimates from procurement dict
    monthly = material_data.get('monthly_forecast', [])
    next_month_qty = float(monthly[0]['forecast_qty']) if monthly else 0
    next_qtr_qty   = sum(m['forecast_qty'] for m in monthly[:3]) if len(monthly) >= 3 else 0
    next_yr_qty    = sum(m['forecast_qty'] for m in monthly[:12]) if len(monthly) >= 12 else 0

    estimates = [
        {'period': 'Next Month',   'quantity': next_month_qty,
         'cost': procurement.get('next_month_cost', 0),   'note': 'Immediate procurement window'},
        {'period': 'Next Quarter', 'quantity': next_qtr_qty,
         'cost': procurement.get('next_quarter_cost', 0), 'note': 'Q1 forward planning estimate'},
        {'period': 'Next Year',    'quantity': next_yr_qty,
         'cost': procurement.get('next_year_cost', 0),    'note': 'Annual budget projection'},
    ]

    for est in estimates:
        bg = COLORS['slate_bg'] if row % 2 == 0 else COLORS['white']
        for col, val in enumerate([
            est.get('period', ''),
            f"{est.get('quantity', 0):,.1f}",
            f"₹{est.get('cost', 0):,.0f}",
            est.get('note', ''),
        ], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.border = BORDER
            c.alignment = Alignment(horizontal='right' if col in (2, 3) else 'left', wrap_text=True)
        row += 1

    row += 1
    _section_title(ws, row, COLS, 'PROCUREMENT INSIGHT', '💡',
                   bg='374151', fg=COLORS['white'])
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    insight_cell = ws.cell(row=row, column=1,
                           value=orders.get('basis', material_data.get('recommendation', '')))
    insight_cell.font = Font(size=10, color=COLORS['navy'])
    insight_cell.fill = PatternFill('solid', fgColor=COLORS['sky'])
    insight_cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[row].height = 48


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 4: Risk Analysis
# ─────────────────────────────────────────────────────────────────────────────

def _write_risk_sheet(wb, material_data: dict):
    ws = wb.create_sheet('Risk Analysis')
    ws.sheet_view.showGridLines = False
    COLS = 4
    _write_title_block(ws, material_data, cols=COLS)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 36
    ws.column_dimensions['D'].width = 20

    row = 4
    _section_title(ws, row, COLS, 'RISK ANALYSIS', '⚠️')
    row += 1

    pct = material_data.get('pct_change', 0)
    status = material_data.get('forecast_status', 'NORMAL')
    health = material_data.get('health', {})
    score = health.get('score', 70)
    lead_time = material_data.get('lead_time', 30)
    stability = material_data.get('demand_stability', 'Unknown')

    # Compute risk flags
    risks = []

    # Demand spike
    if abs(pct) > 25:
        risks.append({
            'risk': 'Demand Spike / Drop Alert',
            'level': 'HIGH' if abs(pct) > 40 else 'MEDIUM',
            'detail': f'Forecast demand change: {pct:+.1f}%. ' + ('Significant demand surge detected.' if pct > 0 else 'Significant demand decline detected.'),
            'action': 'Review procurement schedule immediately',
        })

    # Procurement risk from lead time
    if lead_time > 60:
        risks.append({
            'risk': 'High Lead Time Procurement Risk',
            'level': 'HIGH',
            'detail': f'Lead time is {lead_time:.0f} days, creating supply chain exposure.',
            'action': 'Increase safety stock and consider alternate suppliers',
        })
    elif lead_time > 30:
        risks.append({
            'risk': 'Elevated Lead Time',
            'level': 'MEDIUM',
            'detail': f'Lead time is {lead_time:.0f} days. Monitor supplier performance.',
            'action': 'Review supplier agreements and stock buffer',
        })

    # Inventory instability
    if stability == 'Volatile':
        risks.append({
            'risk': 'Inventory Instability',
            'level': 'HIGH',
            'detail': 'Demand pattern is highly volatile (CV > 35%). Forecasts carry higher uncertainty.',
            'action': 'Implement safety stock review and increase monitoring frequency',
        })
    elif stability == 'Moderate':
        risks.append({
            'risk': 'Moderate Demand Variability',
            'level': 'LOW',
            'detail': 'Demand shows moderate variability. Standard buffers may be adequate.',
            'action': 'Continue monitoring; review if variability increases',
        })

    # Forecast anomaly / low health
    if score < 50:
        risks.append({
            'risk': 'Forecast Anomaly / Low Health',
            'level': 'HIGH',
            'detail': f'Material health score is {score}/100. Multiple risk factors detected.',
            'action': 'Escalate to procurement manager for immediate review',
        })
    elif score < 70:
        risks.append({
            'risk': 'Below-Average Material Health',
            'level': 'MEDIUM',
            'detail': f'Material health score is {score}/100. Some risk factors require attention.',
            'action': 'Schedule procurement review in next planning cycle',
        })

    if not risks:
        risks.append({
            'risk': 'No Significant Risk Detected',
            'level': 'LOW',
            'detail': 'All risk indicators are within acceptable thresholds.',
            'action': 'Maintain standard monitoring cadence',
        })

    # Risk table
    _col_header(ws, row, ['Risk Factor', 'Level', 'Detail', 'Recommended Action'])
    row += 1

    level_bg  = {'HIGH': COLORS['red_bg'],   'MEDIUM': COLORS['amber_bg'], 'LOW': COLORS['green_bg']}
    level_fg  = {'HIGH': COLORS['red_fg'],   'MEDIUM': COLORS['amber_fg'], 'LOW': COLORS['green_fg']}

    for risk in risks:
        lvl = risk['level']
        vals = [risk['risk'], lvl, risk['detail'], risk['action']]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=val)
            if col == 2:
                c.fill = PatternFill('solid', fgColor=level_bg.get(lvl, COLORS['sky']))
                c.font = Font(bold=True, size=9, color=level_fg.get(lvl, COLORS['navy']))
            else:
                c.fill = PatternFill('solid', fgColor=COLORS['slate_bg'] if row % 2 == 0 else COLORS['white'])
                c.font = Font(size=9)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top',
                                    horizontal='center' if col == 2 else 'left')
        ws.row_dimensions[row].height = 42
        row += 1

    # Risk Summary box
    row += 1
    _section_title(ws, row, COLS, 'OVERALL RISK ASSESSMENT', '🎯',
                   bg='374151', fg=COLORS['white'])
    row += 1
    high_count = sum(1 for r in risks if r['level'] == 'HIGH')
    med_count  = sum(1 for r in risks if r['level'] == 'MEDIUM')

    if high_count > 0:
        overall = f'HIGH RISK – {high_count} high-severity issue(s) require immediate attention.'
        overall_bg = COLORS['red_bg']
        overall_fg = COLORS['red_fg']
    elif med_count > 0:
        overall = f'MEDIUM RISK – {med_count} medium-severity issue(s) require monitoring.'
        overall_bg = COLORS['amber_bg']
        overall_fg = COLORS['amber_fg']
    else:
        overall = 'LOW RISK – Material is within acceptable risk thresholds.'
        overall_bg = COLORS['green_bg']
        overall_fg = COLORS['green_fg']

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COLS)
    oc = ws.cell(row=row, column=1, value=overall)
    oc.font = Font(bold=True, size=11, color=overall_fg)
    oc.fill = PatternFill('solid', fgColor=overall_bg)
    oc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Historical & Forecast Charts
# ─────────────────────────────────────────────────────────────────────────────

def _write_charts_sheet(wb, material_data: dict):
    ws = wb.create_sheet('Charts & Trends')
    ws.sheet_view.showGridLines = False
    COLS = 6
    _write_title_block(ws, material_data, cols=COLS)

    chart_data = material_data.get('chart_data', {})
    row = 4

    _section_title(ws, row, COLS, 'HISTORICAL DEMAND & FORECAST CHARTS', '📊')
    row += 2

    charts_embedded = 0
    for key, title in [
        ('line_chart_b64',     'Historical Demand Trend'),
        ('forecast_chart_b64', 'Forecast Trend (Next 12 Months)'),
        ('bar_chart_b64',      'Monthly Demand Distribution'),
    ]:
        b64 = chart_data.get(key)
        if b64:
            try:
                img_bytes = base64.b64decode(b64)
                img_stream = BytesIO(img_bytes)
                img = XLImage(img_stream)
                img.width  = 560
                img.height = 280
                anchor_col = 'A' if charts_embedded % 2 == 0 else 'E'
                anchor_row = row + (charts_embedded // 1) * 17
                # Label row
                lc = ws.cell(row=anchor_row - 1, column=1, value=title)
                lc.font = Font(bold=True, size=10, color=COLORS['navy'])
                ws.add_image(img, f'{anchor_col}{anchor_row}')
                for r in range(anchor_row, anchor_row + 16):
                    ws.row_dimensions[r].height = 18
                row = anchor_row + 17
                charts_embedded += 1
            except Exception:
                pass

    if charts_embedded == 0:
        ws.cell(row=row, column=1,
                value='Charts will be generated when matplotlib is available.').font = \
            Font(size=9, color=COLORS['slate'], italic=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main: Generate Material Report
# ─────────────────────────────────────────────────────────────────────────────

def generate_material_report(material_data: dict) -> BytesIO:
    """
    Build a full professional Excel report for one material.
    Input: material_data dict from get_material_intelligence().
    Returns: BytesIO of the .xlsx workbook.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    _write_overview_sheet(wb, material_data)
    _write_forecast_sheet(wb, material_data)
    _write_procurement_sheet(wb, material_data)
    _write_risk_sheet(wb, material_data)
    _write_charts_sheet(wb, material_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
