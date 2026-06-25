"""
Report Export Module (Phase 1+2+3+4)
Phase 4 adds: Executive Summary sheet, Alerts & Risks sheet, Forecast Notifications sheet
"""
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Optional


# ─── Color palette ───
COLORS = {
    'header_bg': '1E3A5F',
    'header_fg': 'FFFFFF',
    'subheader': '2563EB',
    'peak': 'DCFCE7',
    'peak_text': '166534',
    'normal': 'DBEAFE',
    'normal_text': '1E40AF',
    'low': 'FEE2E2',
    'low_text': '991B1B',
    'alt_row': 'F8FAFC',
    'border': 'CBD5E1',
    'A_class': 'FEF3C7',
    'B_class': 'DBEAFE',
    'C_class': 'F0FDF4',
    'critical': 'FEE2E2',
    'medium':   'FEF3C7',
    'low_alert':'DBEAFE',
}

thin = Side(border_style='thin', color=COLORS['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(ws, row, cols, text, color=None):
    bg = color or COLORS['header_bg']
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill('solid', fgColor=bg)
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = Font(bold=True, color=COLORS['header_fg'], size=12)


def _col_header(ws, row: int, headers: list):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = PatternFill('solid', fgColor=COLORS['subheader'])
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ─────────────────────────────────────────────────────────────
# Phase 4: Executive Summary Sheet (enhanced)
# ─────────────────────────────────────────────────────────────

def write_executive_summary_p4(wb, inv_result: dict, forecast_result: Optional[dict],
                                exec_summary: Optional[dict] = None,
                                alert_result: Optional[dict] = None):
    ws = wb.create_sheet('Executive Summary')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 25

    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value='INVENTORY INTELLIGENCE SYSTEM – EXECUTIVE SUMMARY')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color='1E3A5F')
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='EFF6FF')

    row += 1
    ws.cell(row=row, column=1, value=f'Generated: {datetime.now().strftime("%d %b %Y  %H:%M")}')
    ws.cell(row=row, column=1).font = Font(color='64748B', size=9, italic=True)

    # Headline
    if exec_summary:
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value=exec_summary.get('headline', ''))
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color='1E40AF')
        ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='DBEAFE')
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)

        # KPIs
        row += 2
        ws.cell(row=row, column=1, value='KEY PERFORMANCE INDICATORS').font = Font(bold=True, size=10, color='1E3A5F')
        row += 1
        kpis = exec_summary.get('kpis', {})
        kpi_display = [
            ('Total Materials',        kpis.get('total_items', 0)),
            ('Total Inventory Value',  f"₹{kpis.get('total_value', 0):,.0f}"),
            ('Class A (High Priority)',kpis.get('a_items', 0)),
            ('Peak Demand Materials',  kpis.get('peak_materials', 0)),
            ('Next Quarter Cost Est.', f"₹{kpis.get('next_qtr_cost', 0):,.0f}"),
            ('Cost Change Forecast',   f"{kpis.get('cost_change_pct', 0):+.1f}%"),
            ('Active Alerts',          kpis.get('active_alerts', 0)),
            ('Critical Alerts',        kpis.get('critical_alerts', 0)),
        ]
        for label, val in kpi_display:
            ws.cell(row=row, column=1, value=label).font = Font(size=9, bold=True)
            ws.cell(row=row, column=2, value=str(val)).font = Font(size=9)
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2).border = BORDER
            row += 1

        # Business Summary bullets
        row += 1
        ws.cell(row=row, column=1, value='BUSINESS SUMMARY').font = Font(bold=True, size=10, color='1E3A5F')
        row += 1
        for bullet in exec_summary.get('bullets', []):
            cell = ws.cell(row=row, column=1, value=f'• {bullet}')
            cell.font = Font(size=9)
            cell.fill = PatternFill('solid', fgColor='F8FAFC' if row % 2 else 'FFFFFF')
            ws.merge_cells(f'A{row}:B{row}')
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1

        # Risk flags
        if exec_summary.get('risk_flags'):
            row += 1
            ws.cell(row=row, column=1, value='⚠ RISK FLAGS').font = Font(bold=True, size=10, color='991B1B')
            row += 1
            for flag in exec_summary.get('risk_flags', []):
                cell = ws.cell(row=row, column=1, value=f'⚠ {flag}')
                cell.font = Font(size=9, color='991B1B')
                cell.fill = PatternFill('solid', fgColor='FEE2E2')
                ws.merge_cells(f'A{row}:B{row}')
                cell.alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 30
                row += 1

        # Recommendations
        row += 1
        ws.cell(row=row, column=1, value='✓ RECOMMENDED ACTIONS').font = Font(bold=True, size=10, color='166534')
        row += 1
        for rec in exec_summary.get('recommendations', []):
            cell = ws.cell(row=row, column=1, value=f'✓ {rec}')
            cell.font = Font(size=9, color='166534')
            cell.fill = PatternFill('solid', fgColor='F0FDF4')
            ws.merge_cells(f'A{row}:B{row}')
            cell.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1
    else:
        # Fallback to legacy format
        row += 2
        ws.cell(row=row, column=1, value='📦  Inventory Intelligence Insights').font = Font(bold=True, size=11, color='1E3A5F')
        row += 1
        for insight in inv_result.get('insights', []):
            ws.cell(row=row, column=1, value=f'  • {insight}').font = Font(size=9)
            row += 1

        if forecast_result:
            row += 1
            ws.cell(row=row, column=1, value='📈  Forecast & Demand Intelligence Insights').font = Font(bold=True, size=11, color='1E3A5F')
            row += 1
            for insight in forecast_result.get('exec_insights', []):
                ws.cell(row=row, column=1, value=f'  • {insight}').font = Font(size=9)
                row += 1


# ─────────────────────────────────────────────────────────────
# Phase 4: Alerts & Risks Sheet
# ─────────────────────────────────────────────────────────────

def write_alerts_sheet(wb, alert_result: Optional[dict]):
    ws = wb.create_sheet('Alerts & Risks')
    ws.sheet_view.showGridLines = False

    if not alert_result or not alert_result.get('alerts'):
        ws['A1'] = 'No active alerts at this time.'
        return

    _header_style(ws, 1, 5, 'ALERTS & RISKS – PHASE 4 PRIORITY ENGINE')
    _col_header(ws, 2, ['Priority', 'Category', 'Alert Title', 'Message', 'Affected Materials'])
    ws.row_dimensions[2].height = 30

    priority_fill = {
        'CRITICAL': COLORS['critical'],
        'MEDIUM':   COLORS['medium'],
        'LOW':      COLORS['low_alert'],
    }
    priority_font = {
        'CRITICAL': '991B1B',
        'MEDIUM':   '92400E',
        'LOW':      '1E40AF',
    }

    for i, alert in enumerate(alert_result.get('alerts', []), start=3):
        p = alert.get('priority', 'LOW')
        row_fill = priority_fill.get(p, 'FFFFFF')
        row_data = [
            p,
            alert.get('category', '').replace('_', ' ').title(),
            alert.get('title', ''),
            alert.get('message', ''),
            ', '.join(alert.get('materials', [])[:5]),
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill('solid', fgColor=row_fill)
            cell.border = BORDER
            cell.font = Font(size=9,
                             bold=(j == 1),
                             color=priority_font.get(p, '000000') if j == 1 else '334155')
            cell.alignment = Alignment(horizontal='center' if j == 1 else 'left', wrap_text=True)
        ws.row_dimensions[i].height = 40

    # Summary row
    row = len(alert_result.get('alerts', [])) + 4
    ws.cell(row=row, column=1, value=f"Total: {alert_result.get('total',0)} alerts  |  "
            f"Critical: {alert_result.get('critical_count',0)}  |  "
            f"Medium: {alert_result.get('medium_count',0)}  |  "
            f"Low: {alert_result.get('low_count',0)}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=9, color='1E3A5F')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 30
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# Phase 4: Forecast Notifications Sheet
# ─────────────────────────────────────────────────────────────

def write_forecast_notifications(wb, forecast_result: Optional[dict], alert_result: Optional[dict]):
    ws = wb.create_sheet('Forecast Notifications')
    ws.sheet_view.showGridLines = False

    _header_style(ws, 1, 4, 'FORECAST NOTIFICATIONS – DEMAND INTELLIGENCE ALERTS')
    _col_header(ws, 2, ['Material Code', 'Forecast Status', 'Demand Change %', 'Business Insight'])
    ws.row_dimensions[2].height = 30

    row_num = 3
    if forecast_result:
        forecasts = forecast_result.get('forecasts', [])
        # Show PEAK first, then LOW (actionable ones)
        ordered = ([f for f in forecasts if f.get('forecast_status') == 'PEAK'] +
                   [f for f in forecasts if f.get('forecast_status') == 'LOW'] +
                   [f for f in forecasts if f.get('forecast_status') == 'NORMAL'])
        status_fill = {'PEAK': COLORS['peak'], 'NORMAL': COLORS['normal'], 'LOW': COLORS['low']}
        status_font = {'PEAK': COLORS['peak_text'], 'NORMAL': COLORS['normal_text'], 'LOW': COLORS['low_text']}
        for f in ordered:
            status = f.get('forecast_status', 'NORMAL')
            row_data = [
                f.get('material_code', ''),
                status,
                f"{f.get('pct_change', 0):+.1f}%",
                f.get('insight', ''),
            ]
            for j, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=j, value=val)
                cell.fill = PatternFill('solid', fgColor=status_fill.get(status, 'FFFFFF'))
                cell.border = BORDER
                cell.font = Font(size=9,
                                 bold=(j == 2),
                                 color=status_font.get(status, '000000') if j == 2 else '334155')
                cell.alignment = Alignment(wrap_text=(j == 4))
            ws.row_dimensions[row_num].height = 28
            row_num += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 60
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# Phase 2: Inventory Intelligence Sheet
# ─────────────────────────────────────────────────────────────

def write_inventory_sheet(wb, inv_result: dict):
    ws = wb.create_sheet('Inventory Intelligence')
    ws.sheet_view.showGridLines = False
    df = inv_result.get('data', pd.DataFrame())
    if df.empty:
        ws['A1'] = 'No data available'
        return
    headers = ['Material Code', 'Description', 'Quantity', 'Unit Price (₹)',
               'Annual Value (₹)', 'EOQ', 'Safety Stock', 'Reorder Point',
               'ABC Class', 'Recommendation']
    cols = ['material_code', 'description', 'quantity', 'unit_price',
            'annual_value', 'eoq', 'safety_stock', 'reorder_point',
            'abc_class', 'recommendation']
    _header_style(ws, 1, len(headers), 'INVENTORY INTELLIGENCE – PHASE 2')
    _col_header(ws, 2, headers)
    ws.row_dimensions[2].height = 30
    abc_fill = {'A': COLORS['A_class'], 'B': COLORS['B_class'], 'C': COLORS['C_class']}
    for i, (_, row) in enumerate(df.iterrows(), start=3):
        fill_color = abc_fill.get(str(row.get('abc_class', 'C')), 'FFFFFF')
        for j, col in enumerate(cols, 1):
            val = row.get(col, '')
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill('solid', fgColor=fill_color if i % 2 == 0 else 'FFFFFF')
            cell.border = BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='right' if j > 2 else 'left')
    _auto_width(ws)
    ws.freeze_panes = 'A3'


# ─────────────────────────────────────────────────────────────
# Phase 3: Forecast Sheets
# ─────────────────────────────────────────────────────────────

def write_forecast_summary(wb, forecast_result: dict):
    ws = wb.create_sheet('Forecast Summary')
    ws.sheet_view.showGridLines = False
    forecasts = forecast_result.get('forecasts', [])
    if not forecasts:
        ws['A1'] = 'No forecast data available'
        return
    headers = ['Material Code', 'Description', 'Best Model', 'RMSE', 'MAE', 'R²',
               'Next Month Qty', 'Next Quarter Qty', 'Next Year Qty',
               'Next Month Cost (₹)', 'Next Quarter Cost (₹)', 'Next Year Cost (₹)',
               'Demand Change %', 'Forecast Status', 'Insight']
    _header_style(ws, 1, len(headers), 'FORECAST SUMMARY – PHASE 3')
    _col_header(ws, 2, headers)
    ws.row_dimensions[2].height = 30
    status_fill  = {'PEAK': COLORS['peak'], 'NORMAL': COLORS['normal'], 'LOW': COLORS['low']}
    status_fc    = {'PEAK': COLORS['peak_text'], 'NORMAL': COLORS['normal_text'], 'LOW': COLORS['low_text']}
    for i, f in enumerate(forecasts, start=3):
        m = f.get('model_metrics', {})
        status = f.get('forecast_status', 'NORMAL')
        row_fill = status_fill.get(status, 'FFFFFF')
        row_data = [f.get('material_code',''), f.get('description',''),
                    f.get('best_model','').replace('_',' ').title(),
                    round(m.get('rmse',0),2), round(m.get('mae',0),2), round(m.get('r2',0),3),
                    f.get('next_month_qty',0), f.get('next_quarter_qty',0), f.get('next_year_qty',0),
                    f.get('next_month_cost',0), f.get('next_quarter_cost',0), f.get('next_year_cost',0),
                    f.get('pct_change',0), status, f.get('insight','')]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.fill = PatternFill('solid', fgColor=row_fill)
            cell.border = BORDER
            cell.font = Font(size=9, color=status_fc.get(status,'000000') if j==14 else '000000', bold=(j==14))
            cell.alignment = Alignment(horizontal='right' if j>2 else 'left', wrap_text=(j==15))
    _auto_width(ws)
    ws.freeze_panes = 'A3'


def write_monthly_forecast(wb, forecast_result: dict):
    ws = wb.create_sheet('Monthly Forecast')
    ws.sheet_view.showGridLines = False
    forecasts = forecast_result.get('forecasts', [])
    if not forecasts:
        ws['A1'] = 'No monthly forecast data'
        return
    _header_style(ws, 1, 4, 'MONTHLY FORECAST DETAIL')
    _col_header(ws, 2, ['Material Code', 'Month', 'Forecast Qty', 'Forecast Cost (₹)'])
    ws.row_dimensions[2].height = 25
    row_num = 3
    for f in forecasts:
        for m in f.get('monthly_forecast', []):
            ws.cell(row=row_num, column=1, value=f['material_code']).font = Font(size=9)
            ws.cell(row=row_num, column=2, value=m['month']).font = Font(size=9)
            ws.cell(row=row_num, column=3, value=m['forecast_qty']).font = Font(size=9)
            ws.cell(row=row_num, column=4, value=m['forecast_cost']).font = Font(size=9)
            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border = BORDER
                ws.cell(row=row_num, column=c).fill = PatternFill('solid', fgColor='F8FAFC' if row_num%2==0 else 'FFFFFF')
                ws.cell(row=row_num, column=c).alignment = Alignment(horizontal='right' if c>2 else 'left')
            row_num += 1
    _auto_width(ws)
    ws.freeze_panes = 'A3'


def write_forecast_intelligence(wb, forecast_result: dict):
    ws = wb.create_sheet('Forecast Intelligence')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 70
    row = 1
    ws.cell(row=row, column=1, value='FORECAST INTELLIGENCE – AI-GENERATED INSIGHTS').font = Font(bold=True, size=13, color='1E3A5F')
    row += 2
    for insight in forecast_result.get('exec_insights', []):
        cell = ws.cell(row=row, column=1, value=insight)
        cell.font = Font(size=10)
        cell.fill = PatternFill('solid', fgColor='EFF6FF' if row%2==0 else 'FFFFFF')
        row += 1
    row += 2
    ws.cell(row=row, column=1, value='MATERIAL-LEVEL FORECAST INSIGHTS').font = Font(bold=True, size=11, color='1E3A5F')
    row += 1
    for f in forecast_result.get('forecasts', []):
        status = f.get('forecast_status', 'NORMAL')
        icon = {'PEAK': '🔺', 'NORMAL': '→', 'LOW': '🔻'}.get(status, '→')
        text = f"{icon} [{f['material_code']}] {f.get('insight', '')}"
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(size=9, color={'PEAK': COLORS['peak_text'], 'NORMAL': COLORS['normal_text'], 'LOW': COLORS['low_text']}.get(status, '000000'))
        row += 1


# ─────────────────────────────────────────────────────────────
# Main Export Function (Phase 4 updated signature)
# ─────────────────────────────────────────────────────────────

def generate_excel_report(inv_result: dict,
                          forecast_result: Optional[dict] = None,
                          alert_result: Optional[dict] = None,
                          exec_summary: Optional[dict] = None) -> BytesIO:
    """
    Build full Excel report including Phase 4 sheets:
      - Executive Summary (enhanced)
      - Alerts & Risks (NEW)
      - Forecast Notifications (NEW)
      - Inventory Intelligence
      - Forecast Summary
      - Monthly Forecast
      - Forecast Intelligence
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Phase 4 sheets first (executive view)
    write_executive_summary_p4(wb, inv_result, forecast_result, exec_summary, alert_result)

    if alert_result:
        write_alerts_sheet(wb, alert_result)

    if forecast_result:
        write_forecast_notifications(wb, forecast_result, alert_result)

    # Phase 2 & 3 detail sheets
    write_inventory_sheet(wb, inv_result)
    if forecast_result:
        write_forecast_summary(wb, forecast_result)
        write_monthly_forecast(wb, forecast_result)
        write_forecast_intelligence(wb, forecast_result)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
